import openpyxl

water_sales_workbook = openpyxl.load_workbook("watersales.xlsx")

water_sales_worksheet = water_sales_workbook["Sheet1"]

for row in range(2, water_sales_worksheet.max_row + 2):
    tariff_price = water_sales_worksheet.cell(row=row, column=6).value
    water_sales_202302 = water_sales_worksheet.cell(row=row, column=11).value

    if tariff_price is None and water_sales_202302 is None:
        continue

    result = tariff_price * water_sales_202302

    water_sales_worksheet.cell(row=row, column=12).value = result

water_sales_worksheet.cell(row=1, column=12).value = "February 2023 Billing"

water_sales_workbook.save("watersales_1.xlsx")

