import openpyxl

from main import row

water_sales_workbook = openpyxl.load_workbook("watersales_1.xlsx")
consumption_list = water_sales_workbook["Sheet1"]


for row in range(2, consumption_list.max_row + 1):
    sales_202212 = consumption_list.cell(row=row, column=9).value
    sales_202301 = consumption_list.cell(row=row, column=10).value
    sales_202302 = consumption_list.cell(row=row, column=11).value

    if sales_202212 is None and sales_202301 is None and sales_202302 is None:
        continue

    total_3_months_cons = sales_202212 + sales_202301 + sales_202302

    consumption_list.cell(row=row, column=13).value = total_3_months_cons
    consumption_list.cell(row=1, column=13).value = "Total Cons 3 Months"

water_sales_workbook.save("watersales_2.xlsx")

