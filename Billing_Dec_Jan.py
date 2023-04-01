import openpyxl

workbook = openpyxl.load_workbook("3 months Avg.xlsx")
worksheet = workbook["Sheet1"]

for row in range(2, worksheet.max_row + 1):
    dec_billing = worksheet.cell(row=row, column=9).value
    jan_billing = worksheet.cell(row=row, column=10).value
    tariff_price = worksheet.cell(row=row, column=6).value

    if (dec_billing, jan_billing, tariff_price) is None:
        continue

    dec_billing_1 = dec_billing * tariff_price
    jan_billing_1 = jan_billing * tariff_price

    worksheet.cell(row=row, column=15).value = dec_billing_1
    worksheet.cell(row=1, column=15).value = "December 2022 Billing"

    worksheet.cell(row=row, column=16).value = jan_billing_1
    worksheet.cell(row=1, column=16).value = "January 2023 Billing"

print("Dec 2022 Billing is", dec_billing_1)
print("Jan 2023 Billing is", jan_billing_1)



