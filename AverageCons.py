import openpyxl

average_cons_workbook = openpyxl.load_workbook("watersales_2.xlsx")
average_cons_sheet = average_cons_workbook["Sheet1"]

for row in range(2, average_cons_sheet.max_row + 1):
    sales_202212 = average_cons_workbook.cell(row=row, column=9).value
    sales_202301 = average_cons_workbook.cell(row=row, column=10).value
    sales_202302 = average_cons_workbook.cell(row=row, column=11).value

    if sales_202212 is None and sales_202301 is None and sales_202302 is None:
        continue

    three_months_average = (sales_202212 + sales_202301 + sales_202302) / 3

    average_cons_sheet.cell(row=row, column=14).value = three_months_average
    average_cons_sheet.cell(row=1, column=14).value = "Avg 3 Months Cons"


average_cons_workbook.save("Avg 3 Months.xlsx")
